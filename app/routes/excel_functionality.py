from flask import Blueprint, request, jsonify, current_app, after_this_request, send_file
from flask_jwt_extended import get_current_user, jwt_required
from app.helpers.utils import get_project_id, create_slug, has_access_to_project
from app.models.PayloadModel import PayloadModel
from app.models.ExpectedOutcomeModel import ExpectedOutcomeModel
from app.models.TeststepModel import TestStepModel
from app.models.TestdataModel import TestdataModel, TestdataSchema
from openpyxl import load_workbook
import pandas as pd
import json
import os
import re

testdata_excel_blueprint = Blueprint('testdata_excel', __name__)
email_regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
datetime_regex = r'/^(0[1-9]|1\d|2\d|3[01])\-(0[1-9]|1\d|2\d|3[01])\-(19|20)\d{2}$/'
date_regex = r'^([0-2][0-9]|(3)[0-1])(\/)(((0)[0-9])|((1)[0-2]))(\/)\d{4}$'
number_regex = r'^[-+]?[0-9]+$'


@testdata_excel_blueprint.route("/download", methods=["GET"])
@jwt_required()
def testdata_excel_downloader():
    try:
        user = get_current_user().id
        project = get_project_id(request.args.get('project'))
        teststep = request.args.get('teststep')
        if has_access_to_project(project_id=project, user_id=user):
            teststep_object = TestStepModel.get_one_teststep(id=teststep)
            if teststep_object:
                testdata = TestdataModel.get_all_testdatas(teststep_id=teststep)
                payload_template = PayloadModel.get_one_payload(teststep_object.get('payload_id'))
                expected_outcome_templates = ExpectedOutcomeModel.get_all_expected_outcomes(payload_id=teststep_object.get('payload_id'))
                file_name = json_to_excel(payload=payload_template, testdata=testdata, expected_outcome=expected_outcome_templates,user=user, project=request.args.get('project'))
                @after_this_request
                def remove_file(response):
                    file_name1 = os.path.join(current_app.config['DOWNLOAD_FOLDER'], file_name)
                    os.remove(path=file_name1)
                    return response
                return send_file(os.path.join("../"+current_app.config['DOWNLOAD_FOLDER'], file_name), as_attachment=True)
        else:
            return jsonify({"error": "You do not have access to this project, kindly request the project admin to get access to this project"}), 401
    except Exception as err:
        print(str((err)))
        return jsonify({"error": "something went wrong"}), 400


@testdata_excel_blueprint.route("/upload", methods=["POST"])
@jwt_required()
def testdata_uploader():
    """
    This route takes excel sheet uploaded by the user and uploads the excel data into the database
    """
    try:
        user = get_current_user().id
        project = get_project_id(request.form['project'])
        if has_access_to_project(project_id=project, user_id=user):
            data = request.files['file']
            data.save(os.path.join(current_app.config['UPLOAD_FOLDER'], data.filename))
            sheet_data = pd.read_excel(str(os.path.join(current_app.config['UPLOAD_FOLDER']))+"/"+str(data.filename), sheet_name="Testdata_Combination", skiprows=[0, 1])
            sheet_data = sheet_data.to_json(orient="records")
            sheet_data = json.loads(sheet_data)
            testdata_range, expected_outcome_range = field_range_counter(excel_sheet=data)
            formatted_json_testdata = testdata_json_formatting(sheet_data, testdata_range, expected_outcome_range)
            is_json_data_valid = testdata_validor(formatted_json_testdata, user, project)
            os.remove((os.path.join(current_app.config['UPLOAD_FOLDER']))+"/"+str(data.filename))
            if is_json_data_valid == True:
                formatted_json_testdata = expected_outcome_modifier(testdata_json=formatted_json_testdata)
                testdata_json_to_models(formatted_json_testdata, project, user)
                return jsonify({"message": "testdata updated successfully"}), 200
            else:
                return jsonify({"error": is_json_data_valid["error"]}), 400
        else:
            return jsonify({"error": "You do not have access to this project,kindly connect with the admin to make interact with the project components"}), 401
    except Exception as err:
        print(str(err))
        return jsonify({"error": "something went wrong"}), 400


def expected_outcome_modifier(testdata_json):
    try:
        for testdata in testdata_json:
            outcome = []
            expected_outcome = testdata["Expected_Outcome"]
            for key, value in expected_outcome.items():
                if value is not None:
                    if (re.fullmatch(number_regex, str(value))):
                        Type = "number"
                    elif (re.fullmatch(email_regex, str(value))):
                        Type = "email"
                    elif (re.fullmatch(datetime_regex, str(value))) or (re.fullmatch(date_regex, str(value))):
                        Type = "dateTime"
                    else:
                        Type = "text"
                else:
                    Type = "Null"
                outcome.append({"name": key, "value": value,"type": Type, "isExact": True})
            testdata["Expected_Outcome"] = outcome
        return testdata_json
    except Exception as err:
        return {"error": str(err)}


def field_range_counter(excel_sheet):
    testdata_range = {"start": 0, "end": 0, "range": 0}
    expected_outcome_range = {"start": 0, "end": 0, "range": 0}
    testdata_count_completed = False
    sheet_data = pd.read_excel(str(os.path.join(current_app.config['UPLOAD_FOLDER']))+"/"+str(
        excel_sheet.filename), sheet_name="Testdata_Combination")
    counter = 0
    for cell in sheet_data.loc[0, ]:
        if cell == "Testdata" and testdata_range['start'] == 0:
            testdata_range['start'] = counter
            testdata_range['range'] += 1
        elif cell == "Expected_Outcome":
            testdata_count_completed = True
            testdata_range['end'] = counter
            expected_outcome_range['start'] = counter
            expected_outcome_range['range'] += 1
        elif testdata_range['start'] > 0 and testdata_count_completed == False:
            testdata_range['range'] += 1
        elif testdata_count_completed:
            expected_outcome_range['range'] += 1
        counter += 1
    expected_outcome_range["end"] = counter
    return testdata_range, expected_outcome_range


def testdata_json_formatting(sheet_data, testdata_range, expected_outcome_range):
    combined_json_data = []
    last_row_pointer = 0
    loop_counter = 0
    for row_data in sheet_data:
        jsn = {"Testdata": {}, "Expected_Outcome": {}}
        if row_data["Index"] == 1:
            itr = 0
            for key, value in row_data.items():
                if itr < testdata_range['start']:
                    if key == "Teststep Name":
                        if value is not None:
                            jsn[key] = create_slug(value)
                        else:
                            jsn[key] = value
                    elif key == "Testdata Combination Name":
                        if value is not None:
                            jsn[key] = create_slug(value)
                        else:
                            jsn[key] = value
                if itr > testdata_range['start'] and itr < testdata_range['end']:
                    if value != "skip":
                        jsn["Testdata"][key] = value
                elif itr >= expected_outcome_range['start']:
                    if value != "skip":
                        if key == "status_code":
                            jsn["Expected_Outcome"][key] = int(value)
                        else:
                            jsn["Expected_Outcome"][key] = value
                itr += 1
            last_row_pointer = loop_counter
            combined_json_data.append(jsn)
        else:
            itr = 0
            for key, value in row_data.items():
                if itr > testdata_range['start'] and itr < testdata_range['end']:
                    if value is not None:
                        if type(combined_json_data[last_row_pointer]["Testdata"][key]) is list:
                            combined_json_data[last_row_pointer]["Testdata"][key].append(
                                value)
                        else:
                            temp_value = combined_json_data[last_row_pointer]["Testdata"][key]
                            combined_json_data[last_row_pointer]["Testdata"][key] = [
                                temp_value, value]
                elif itr >= expected_outcome_range['start']:
                    if value is not None:
                        if type(combined_json_data[last_row_pointer]["Expected_Outcome"][key]) is list:
                            combined_json_data[last_row_pointer]["Expected_Outcome"][key].append(
                                value)
                        else:
                            temp_value = combined_json_data[last_row_pointer]["Expected_Outcome"][key]
                            combined_json_data[last_row_pointer]["Expected_Outcome"][key] = [
                                temp_value, value]
                itr += 1
        loop_counter += 1
    return combined_json_data


def testdata_validor(formatted_json_testdata, user, project):
    try:
        for testdata in formatted_json_testdata:
            teststep = testdata["Teststep Name"]
            is_teststep = TestStepModel.is_exist(
                name=teststep, project=project)
            if not is_teststep:
                return {"error": "Teststep with name " + str(teststep) + "does not exist for this project, failed to save data"}

            if testdata["Expected_Outcome"].get("status_code") == None:
                return {"error": "Teststep name " + str(teststep) + "does not posess status code, it is mandatory to pass status code for the request"}
            else:
                if type(testdata["Expected_Outcome"]["status_code"]) is list:
                    return {"error": "Teststep name " + str(teststep) + " has multiple status codes, kindly remove the extra status codes"}
        return True
    except Exception as err:
        return {"error": str(err)}


def testdata_json_to_models(formatted_json_testdata, project, user):
    for testdata in formatted_json_testdata:
        testdata["teststep"] = TestStepModel.is_exist(
            name=testdata["Teststep Name"], project=project).id
        del testdata["Teststep Name"]
        testdata["payload"] = testdata["Testdata"]
        del testdata["Testdata"]
        testdata["expected_outcome"] = testdata["Expected_Outcome"]
        del testdata["Expected_Outcome"]
        testdata["name"] = testdata["Testdata Combination Name"]
        del testdata["Testdata Combination Name"]
        testdata["parameters"] = {}
        testdata["created_by"] = user
        testdata["modified_by"] = user
        td = TestdataModel.is_exist(testdata['name'], testdata['teststep'])
        if td is not None:
            td = TestdataModel.query.filter_by(
                name=testdata['name'], teststep=testdata['teststep']).first()
            td.payload = testdata['payload']
            td.expected_outcome = testdata['expected_outcome']
            td.created_by = user
            td.modified_by = user
            td.save()
        else:
            testdata = TestdataSchema().load(testdata)
            testdata = TestdataModel(testdata)
            testdata.save()


def expected_outcome_to_json(expected_outcome):
    json_expected_outcome = {}
    for field in expected_outcome:
        key = field['name']
        value = field['value']
        json_expected_outcome[key] = value
    return json_expected_outcome


def json_to_excel(payload, testdata, expected_outcome, user, project):
    combined_json_data = []
    for data in expected_outcome:
        data['expected_outcome'] = expected_outcome_to_json(data["expected_outcome"])
    for data in testdata:
        data["expected_outcome"] = expected_outcome_to_json(data["expected_outcome"])
    testdata_key_set = set()
    expected_outcome_key_set = set()

    for key, value in payload["payload"].items():
        testdata_key_set.add(key)

    for data in testdata:
        for key, value in data["payload"].items():
            testdata_key_set.add(key)

    for data in expected_outcome:
        for key, value in data["expected_outcome"].items():
            expected_outcome_key_set.add(key)

    for data in testdata:
        for key, value in data["expected_outcome"].items():
            expected_outcome_key_set.add(key)

    for data in testdata:
        del data['parameters']
        del data['modified_by']
        del data['created_by']
        del data['modified_at']
        del data['created_at']
        del data['id']
        data['Testdata Combination Name'] = data['name']
        del data['name']
        data['Teststep Name'] = TestStepModel.get_one_teststep(
            id=data['teststep']).get('name')
        del data['teststep']
        combined_json_data.append(data)

    for td in combined_json_data:
        for key in testdata_key_set:
            if key not in td['payload']:
                td["payload"][key] = "skip"

        for key in expected_outcome_key_set:
            if key not in td['expected_outcome']:
                td['expected_outcome'][key] = "skip"

    for ind, td in enumerate(combined_json_data):
        td_new = {
            'ID': ind+1,
            'Teststep Name': td['Teststep Name'],
            'Testdata Combination Name': td['Testdata Combination Name'],
            'Index': 1,
            **td['payload'],
            **td['expected_outcome']
        }
        combined_json_data[ind] = td_new
        del td_new

    final_json_list = []
    for ind, td in enumerate(combined_json_data):
        if has_nested_values(td) is True:
            lst_itr = 0
            are_lists_iterated = False
            while are_lists_iterated is not True:
                td_new = {}
                for key, value in td.items():
                    if lst_itr == 0:
                        if type(value) is not list:
                            td_new[key] = value
                        else:
                            td_new[key] = value[0]
                    else:
                        if type(value) is list and lst_itr < len(value):
                            td_new[key] = value[lst_itr]
                            td_new['Index'] = lst_itr+1
                if len(td_new) > 0:
                    final_json_list.append(td_new)
                else:
                    are_lists_iterated = True
                lst_itr += 1
        else:
            final_json_list.append(td)
    final_json_list = json.dumps(final_json_list)
    df = pd.read_json(final_json_list)
    file_name = "Testdata-Combination-" + str(project) + "-project" + ".xlsx"
    df.to_excel(excel_writer=os.path.join(current_app.config['DOWNLOAD_FOLDER'], file_name), sheet_name="Testdata_Combination", index=False, startrow=2)
    testdata_range = len(testdata_key_set) + 1
    expected_outcome_range = len(expected_outcome_key_set)
    work_book = load_workbook(filename=os.path.join(current_app.config['DOWNLOAD_FOLDER'], file_name))
    work_sheet = work_book.active
    work_sheet['A2'] = 'Testdata fields'
    work_sheet.merge_cells('A2:C2')
    work_sheet['D2'] = 'Testdata'
    work_sheet.merge_cells(start_row=2, start_column=4, end_row=2, end_column=3+testdata_range)
    for row in work_sheet.iter_rows(min_row=2,max_row=2):
        for cell in row:
            if cell.coordinate > 'D2':
                testdata_range -= 1
            if testdata_range == 0:
                cell.value = 'Expected_Outcome'
                break
    work_book.save(filename=os.path.join(current_app.config['DOWNLOAD_FOLDER'], file_name))
    return file_name


def has_nested_values(testdata):
    for key, value in testdata.items():
        if type(value) == list:
            return True

    return False
